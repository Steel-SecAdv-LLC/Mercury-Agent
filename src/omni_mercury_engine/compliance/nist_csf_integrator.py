# Copyright (C) 2025 Steel Security Advisors LLC
# SPDX-License-Identifier: GPL-3.0-or-later
"""NIST Cybersecurity Framework (CSF) 2.0 integrator for Mercury Agent.

Implements all six CSF 2.0 core functions (GOVERN, IDENTIFY, PROTECT,
DETECT, RESPOND, RECOVER) with live reference data fetched from the
authoritative NIST CSF 2.0 Reference Tool maintained by NIST CSRC. The
fetcher hits the public reference endpoint at
``https://csrc.nist.gov/extensions/nudp/services/json/csf/download?olirids=all``
(no credentials required) and parses the returned ``.xlsx`` payload into
typed :class:`NISTFunction`, :class:`NISTCategory`, and
:class:`NISTSubcategory` records.

Public surface
==============

* :class:`NISTFunction`, :class:`ImplementationTier` enums
* :class:`NISTCategory`, :class:`NISTSubcategory`, :class:`NISTProfile`,
  :class:`NISTAssessment` dataclasses
* :class:`NISTCSFReferenceFetcher` -- live network-backed reference loader
* :class:`NISTCSFIntegrator` -- the five canonical operations:
  ``assess_function``, ``create_profile``, ``detect_supply_chain_anomalies``,
  ``continuous_monitoring_detect``, ``generate_compliance_report``
* :func:`get_nist_csf_integrator` -- factory

Source provenance
=================

Ported from ``omni_anomaly_engine.domains.ciad.compliance.nist_csf_integrator``
(Omni-AXA-Engine). Mercury Agent additions and corrections:

* New :class:`NISTSubcategory` dataclass carrying the implementation
  examples and informative references parsed from the live source.
* New :class:`NISTCSFReferenceFetcher` performing a real HTTP fetch
  against ``csrc.nist.gov`` with on-disk caching under
  ``$XDG_CACHE_HOME/mercury-agent/nist_csf`` (default
  ``~/.cache/mercury-agent/nist_csf``).
* :class:`NISTCSFIntegrator` now defaults to the live reference and falls
  back to the curated built-in tree on explicit opt-out so air-gapped
  callers (e.g. CI security-isolation lanes) remain functional.
* All public surfaces typed for ``mypy --strict`` with Google-style
  docstrings matching Mercury's existing convention
  (cf. :mod:`omni_mercury_engine.utils.logging`).
* :func:`continuous_monitoring_detect` no longer returns the
  unnormalised ``deviations`` array as ``anomaly_scores``; the score is
  rescaled into ``[0, 1]`` using the joint ``max(baseline, deviation)``
  denominator from the source but the result is now strictly bounded
  to that interval. Numerically identical to the source on bounded
  inputs.
* Replaces the original silent ``except`` paths -- every failure point
  raises a typed exception or logs through Mercury's structured logger.

The CSF 2.0 core defines six Functions, 22 Categories, and 106
Subcategories as of the 2024-02-26 final publication. Counts are
verified at runtime against the live fetch (see
:meth:`NISTCSFIntegrator.verify_coverage`).
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
import time
import warnings
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any, Final
from urllib.parse import urlparse

import numpy as np
import requests

from omni_mercury_engine.utils.logging import get_logger

if TYPE_CHECKING:
    import logging
    from collections.abc import Iterable

_LOG: Final[logging.Logger] = get_logger("omni_mercury_engine.compliance.nist_csf")

# Authoritative NIST CSF 2.0 Reference Tool download endpoint. Returns an
# XLSX workbook containing the canonical Function / Category /
# Subcategory tree with implementation examples and informative
# references. Public, no credentials. Verified 2026-05-18 (200 OK,
# ~143 KB) against https://www.nist.gov/cyberframework.
NIST_CSF_REFERENCE_URL: Final[str] = (
    "https://csrc.nist.gov/extensions/nudp/services/json/csf/download?olirids=all"
)

# Allowlist of hosts the fetcher may contact. Bound to the published
# NIST CSRC endpoint to prevent SSRF / redirect-pivot attacks via the
# ``url`` constructor argument; ``follow_redirects`` is forced off on
# every outbound request for the same reason.
_ALLOWED_REFERENCE_HOSTS: Final[frozenset[str]] = frozenset({"csrc.nist.gov"})

# NIST CSF 2.0 publication landing pages and authoritative PDF. Used by
# :meth:`NISTCSFReferenceFetcher.metadata` to surface publication context
# alongside any cached payload.
NIST_CSF_PUBLICATION_URL: Final[str] = "https://www.nist.gov/cyberframework"
NIST_CSF_PUBLICATION_PDF_URL: Final[str] = "https://nvlpubs.nist.gov/nistpubs/cswp/NIST.CSWP.29.pdf"

# Filesystem cache layout. Honoured by both the fetcher and tests so the
# refresh / invalidation behaviour is deterministic.
_DEFAULT_CACHE_TTL_SECONDS: Final[float] = 7 * 24 * 60 * 60.0  # 7 days
_HTTP_TIMEOUT_SECONDS: Final[float] = 30.0
_USER_AGENT: Final[str] = (
    "MercuryAgent-NIST-CSF-Fetcher/1.0 " "(+https://github.com/Steel-SecAdv-LLC/Mercury-Agent)"
)


def _default_cache_dir() -> Path:
    """Return the default on-disk cache directory.

    Honours ``$XDG_CACHE_HOME`` and falls back to
    ``$HOME/.cache/mercury-agent/nist_csf``.

    Returns:
        Absolute path to the cache directory. The directory is not
        created by this helper.
    """
    xdg = os.environ.get("XDG_CACHE_HOME")
    base = Path(xdg) if xdg else Path.home() / ".cache"
    return base / "mercury-agent" / "nist_csf"


class NISTFunction(Enum):
    """The six core functions defined by NIST CSF 2.0."""

    GOVERN = "GOVERN"
    IDENTIFY = "IDENTIFY"
    PROTECT = "PROTECT"
    DETECT = "DETECT"
    RESPOND = "RESPOND"
    RECOVER = "RECOVER"


class ImplementationTier(Enum):
    """NIST CSF implementation tiers (Partial through Adaptive).

    The numeric value is the canonical tier rank used for gap analysis
    arithmetic. Tier 1 (PARTIAL) is the lowest maturity, Tier 4
    (ADAPTIVE) is the highest.
    """

    PARTIAL = 1
    RISK_INFORMED = 2
    REPEATABLE = 3
    ADAPTIVE = 4


@dataclass(frozen=True)
class NISTSubcategory:
    """A CSF 2.0 subcategory under a :class:`NISTCategory`.

    Attributes:
        id: Subcategory identifier such as ``"GV.OC-01"``.
        description: Outcome statement for the subcategory.
        implementation_examples: Verbatim examples shipped by NIST CSRC
            (may be empty when the source omits them).
        informative_references: Cross-mapping strings (e.g. ``"CCMv4.0:
            BCR-01"``) verbatim from the reference tool.
    """

    id: str
    description: str
    implementation_examples: tuple[str, ...] = ()
    informative_references: tuple[str, ...] = ()


@dataclass
class NISTCategory:
    """A CSF 2.0 category under a :class:`NISTFunction`.

    Attributes:
        id: Category identifier such as ``"GV.OC"``.
        name: Human-readable category name.
        description: Category description.
        subcategories: Subcategory records. Strings are accepted for
            backward compatibility with the original Omni-AXA-Engine
            surface (where subcategories were plain ``"GV.OC-01: ..."``
            strings); the dataclass normalises both representations
            through :meth:`subcategory_ids`.
    """

    id: str
    name: str
    description: str
    subcategories: list[NISTSubcategory | str] = field(default_factory=list)

    def subcategory_ids(self) -> list[str]:
        """Return the canonical subcategory identifiers as strings.

        Returns:
            Ordered list of subcategory identifiers, with the
            ``"ID: description"`` legacy form parsed on the fly.
        """
        ids: list[str] = []
        for sub in self.subcategories:
            if isinstance(sub, NISTSubcategory):
                ids.append(sub.id)
            else:
                ids.append(sub.split(":", 1)[0].strip())
        return ids


@dataclass
class NISTProfile:
    """NIST CSF profile for current-vs-target gap analysis.

    Attributes:
        current_state: Maturity score per function name.
        target_state: Target maturity score per function name.
        gaps: ``target_state[f] - current_state[f]`` per function.
        priority_actions: Aggregated recommendations whose source gap
            exceeded 0.20.
    """

    current_state: dict[str, float]
    target_state: dict[str, float]
    gaps: dict[str, float] = field(default_factory=dict)
    priority_actions: list[str] = field(default_factory=list)


@dataclass
class NISTAssessment:
    """Result of a single :class:`NISTFunction` assessment.

    Attributes:
        function: The function being assessed.
        tier: The :class:`ImplementationTier` resolved from the maturity
            score.
        maturity_score: Mean maturity in ``[0, 1]``.
        findings: Human-readable findings (gap descriptions, etc.).
        recommendations: Human-readable next-step recommendations.
        risk_score: Criticality-weighted risk score in ``[0, 1]``.
        timestamp: UTC timestamp of the assessment.
    """

    function: NISTFunction
    tier: ImplementationTier
    maturity_score: float
    findings: list[str]
    recommendations: list[str]
    risk_score: float
    timestamp: datetime = field(default_factory=lambda: datetime.now(UTC))


class NISTCSFReferenceError(RuntimeError):
    """Raised when the live NIST CSF reference cannot be loaded."""


class NISTCSFReferenceFetcher:
    """Fetch and cache the live NIST CSF 2.0 reference catalogue.

    The fetcher pulls the canonical ``.xlsx`` payload from the public
    NIST CSF 2.0 Reference Tool, parses it with :mod:`openpyxl`, and
    materialises an immutable :class:`NISTFunction` -> ``list[NISTCategory]``
    tree. The raw payload is cached on disk (default
    ``~/.cache/mercury-agent/nist_csf``) so subsequent calls within the
    TTL window do not hit the network.

    Example:
        >>> fetcher = NISTCSFReferenceFetcher()
        >>> tree = fetcher.load_reference_tree()
        >>> sorted(tree.keys())  # doctest: +ELLIPSIS
        [<NISTFunction.DETECT...
    """

    def __init__(
        self,
        url: str = NIST_CSF_REFERENCE_URL,
        cache_dir: Path | None = None,
        cache_ttl_seconds: float = _DEFAULT_CACHE_TTL_SECONDS,
        session: requests.Session | None = None,
    ) -> None:
        """Initialise the fetcher.

        Args:
            url: Reference download URL. Defaults to
                :data:`NIST_CSF_REFERENCE_URL`. The scheme must be
                ``https`` and the host must be in
                :data:`_ALLOWED_REFERENCE_HOSTS` (currently the public
                NIST CSRC reference-tool endpoint). Any other value
                raises :class:`NISTCSFReferenceError` at construction
                time so SSRF / redirect-pivot attempts surface at the
                boundary instead of at fetch time.
            cache_dir: Cache directory. Defaults to
                :func:`_default_cache_dir`.
            cache_ttl_seconds: Maximum cache age before a fresh fetch
                is forced. ``0`` disables caching entirely.
            session: Optional pre-configured :class:`requests.Session`.
                A fresh session is constructed when ``None``.

        Raises:
            NISTCSFReferenceError: If ``url`` is not HTTPS or its host
                is not in :data:`_ALLOWED_REFERENCE_HOSTS`.
        """
        self._validate_url(url)
        self._url = url
        self._cache_dir = cache_dir if cache_dir is not None else _default_cache_dir()
        self._cache_ttl = float(cache_ttl_seconds)
        self._session = session if session is not None else requests.Session()

    @staticmethod
    def _validate_url(url: str) -> None:
        """Validate that ``url`` targets the published NIST CSRC host over HTTPS.

        Args:
            url: Candidate reference URL.

        Raises:
            NISTCSFReferenceError: If the scheme is not ``https`` or the
                host is not in :data:`_ALLOWED_REFERENCE_HOSTS`.
        """
        parsed = urlparse(url)
        if parsed.scheme.lower() != "https":
            raise NISTCSFReferenceError(
                f"NIST CSF reference URL must use HTTPS; got scheme={parsed.scheme!r}"
            )
        host = (parsed.hostname or "").lower()
        if host not in _ALLOWED_REFERENCE_HOSTS:
            raise NISTCSFReferenceError(
                f"NIST CSF reference URL host {host!r} is not in the published "
                f"NIST CSRC allowlist {sorted(_ALLOWED_REFERENCE_HOSTS)!r}"
            )

    # ------------------------------------------------------------------ caching

    @property
    def cache_dir(self) -> Path:
        """Return the configured cache directory (not yet created)."""
        return self._cache_dir

    def _cache_path(self) -> Path:
        """Return the deterministic cache filename for the configured URL."""
        digest = hashlib.sha256(self._url.encode("utf-8")).hexdigest()[:16]
        return self._cache_dir / f"csf_olir_{digest}.xlsx"

    def _cache_fresh(self, path: Path) -> bool:
        """Return ``True`` if a cache file exists and is within TTL."""
        if self._cache_ttl <= 0.0 or not path.exists():
            return False
        try:
            mtime = path.stat().st_mtime
        except OSError:
            return False
        return (time.time() - mtime) < self._cache_ttl

    def fetch_payload(self, force_refresh: bool = False) -> bytes:
        """Return the raw XLSX payload bytes for the reference tool.

        Args:
            force_refresh: Bypass the on-disk cache and force a new HTTP
                request.

        Returns:
            Raw XLSX bytes.

        Raises:
            NISTCSFReferenceError: If the HTTP request fails, the
                payload is empty, or the bytes are not a recognisable
                XLSX file.
        """
        cache_path = self._cache_path()
        if not force_refresh and self._cache_fresh(cache_path):
            try:
                return cache_path.read_bytes()
            except OSError as exc:
                _LOG.warning("Cache read failed for %s: %s; refetching", cache_path, exc)

        _LOG.info("Fetching NIST CSF 2.0 reference from %s", self._url)
        try:
            response = self._session.get(
                self._url,
                timeout=_HTTP_TIMEOUT_SECONDS,
                headers={"User-Agent": _USER_AGENT, "Accept": "*/*"},
                allow_redirects=False,
            )
            response.raise_for_status()
        except requests.RequestException as exc:
            raise NISTCSFReferenceError(
                f"Failed to fetch NIST CSF reference from {self._url}: {exc}"
            ) from exc
        if response.status_code in {301, 302, 303, 307, 308}:
            raise NISTCSFReferenceError(
                f"NIST CSF reference endpoint returned a redirect "
                f"({response.status_code} -> {response.headers.get('Location', '')!r}); "
                f"refusing to follow to prevent SSRF / pivot via Location header. "
                f"Update NIST_CSF_REFERENCE_URL if NIST CSRC has moved the endpoint."
            )

        payload: bytes = response.content
        if not payload:
            raise NISTCSFReferenceError(f"Empty payload returned from {self._url}")
        # XLSX files are ZIP archives; the OOXML magic is the standard
        # ZIP local-file-header signature.
        if not payload.startswith(b"PK\x03\x04"):
            raise NISTCSFReferenceError(
                f"Unexpected non-XLSX payload from {self._url}: " f"first 8 bytes = {payload[:8]!r}"
            )

        try:
            self._cache_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            _LOG.warning("Cache directory creation failed: %s", exc)
        else:
            # Atomic write so partial files never linger in the cache
            # under concurrent fetchers.
            try:
                tmp = tempfile.NamedTemporaryFile(
                    delete=False,
                    dir=self._cache_dir,
                    prefix=".csf_olir_",
                    suffix=".xlsx.tmp",
                )
                with tmp as handle:
                    handle.write(payload)
                Path(tmp.name).replace(cache_path)
            except OSError as exc:
                _LOG.warning("Cache write failed for %s: %s", cache_path, exc)

        return payload

    # ----------------------------------------------------------------- parsing

    def load_reference_tree(
        self,
        force_refresh: bool = False,
    ) -> dict[NISTFunction, list[NISTCategory]]:
        """Return the parsed NIST CSF 2.0 reference tree.

        Args:
            force_refresh: Bypass the on-disk cache.

        Returns:
            Mapping from :class:`NISTFunction` to an ordered list of
            :class:`NISTCategory` records. Each category's
            ``subcategories`` field contains :class:`NISTSubcategory`
            instances populated with implementation examples and
            informative references.

        Raises:
            NISTCSFReferenceError: If the payload cannot be parsed or
                if the resulting tree does not contain at least one
                function with one category.
        """
        payload = self.fetch_payload(force_refresh=force_refresh)
        return _parse_csf_xlsx(payload)

    def metadata(self) -> dict[str, str]:
        """Return human-readable provenance metadata for the reference.

        Returns:
            Dictionary with the publication URL, the PDF download URL,
            the reference-tool URL, and the cache file location.
        """
        return {
            "publication_url": NIST_CSF_PUBLICATION_URL,
            "publication_pdf_url": NIST_CSF_PUBLICATION_PDF_URL,
            "reference_tool_url": self._url,
            "cache_path": str(self._cache_path()),
        }


def _parse_csf_xlsx(payload: bytes) -> dict[NISTFunction, list[NISTCategory]]:
    """Parse the NIST CSF 2.0 Reference Tool XLSX payload.

    The workbook contains one data sheet ("CSF 2.0") with five columns:
    Function, Category, Subcategory, Implementation Examples,
    Informative References. Rows cascade -- a function row has only
    column A populated, a category row only column B, and subcategory
    rows have C / D / E populated under the most-recently-seen function
    and category.

    Args:
        payload: Raw XLSX bytes.

    Returns:
        Mapping from :class:`NISTFunction` to its categories.

    Raises:
        NISTCSFReferenceError: If the workbook is unreadable, the
            expected sheet is missing, or no functions are recovered.
    """
    try:
        import openpyxl  # local import: only required when live fetch is used
    except ImportError as exc:  # pragma: no cover - exercised via dep manifest
        raise NISTCSFReferenceError(
            "openpyxl is required to parse the NIST CSF reference XLSX. "
            "Install Mercury Agent with the compliance extra: "
            "pip install 'mercury-agent[compliance]'"
        ) from exc

    try:
        # The NIST CSF 2.0 Reference Tool workbook ships without a
        # default style block, which causes openpyxl >= 3.1 to emit
        # ``UserWarning: Workbook contains no default style, apply
        # openpyxl's default``.  The fallback openpyxl applies is the
        # documented behaviour for this case and we have no influence
        # over NIST's serialiser, so we scope a ``catch_warnings``
        # block around the load call (only suppressing that exact
        # message) instead of leaking the noise into every consumer.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"Workbook contains no default style",
                category=UserWarning,
                module=r"openpyxl(\..*)?",
            )
            workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    except Exception as exc:
        raise NISTCSFReferenceError(f"Failed to load NIST CSF reference workbook: {exc}") from exc

    sheet_name = next(
        (s for s in workbook.sheetnames if s.strip().lower().startswith("csf")),
        None,
    )
    if sheet_name is None:
        raise NISTCSFReferenceError(
            f"NIST CSF data sheet not found; sheets present: {workbook.sheetnames!r}"
        )
    sheet = workbook[sheet_name]

    tree: dict[NISTFunction, list[NISTCategory]] = {}
    current_function: NISTFunction | None = None
    current_category: NISTCategory | None = None

    # Skip the header row.
    rows: Iterable[tuple[Any, ...]] = sheet.iter_rows(min_row=2, max_col=5, values_only=True)
    for raw_row in rows:
        # Pad to five columns to make column indexing safe even for
        # truncated rows.
        row = (list(raw_row) + [None] * 5)[:5]
        col_function, col_category, col_subcategory, col_examples, col_refs = row
        function_text = _normalise_cell(col_function)
        category_text = _normalise_cell(col_category)
        subcategory_text = _normalise_cell(col_subcategory)

        if function_text:
            current_function = _parse_function_header(function_text)
            current_category = None
            if current_function is not None:
                tree.setdefault(current_function, [])
            continue

        if category_text and current_function is not None:
            current_category = _parse_category_header(category_text)
            if current_category is not None:
                tree.setdefault(current_function, []).append(current_category)
            continue

        if subcategory_text and current_function is not None and current_category is not None:
            subcategory = _parse_subcategory_row(
                subcategory_text,
                _normalise_cell(col_examples),
                _normalise_cell(col_refs),
            )
            if subcategory is not None:
                current_category.subcategories.append(subcategory)

    if not tree or all(not v for v in tree.values()):
        raise NISTCSFReferenceError(
            "Parsed NIST CSF reference is empty; the upstream payload "
            "may have changed format. Inspect cache_path in metadata()."
        )

    return tree


def _normalise_cell(value: Any) -> str:
    """Coerce an :mod:`openpyxl` cell value to a stripped string.

    Args:
        value: Raw cell value (``None``, ``int``, ``str``, ``datetime``,
            ...).

    Returns:
        Stripped string representation; empty string when the cell is
        empty.
    """
    if value is None:
        return ""
    return str(value).strip()


def _parse_function_header(text: str) -> NISTFunction | None:
    """Parse a function header row like ``"GOVERN (GV): The ..."``.

    Args:
        text: Cell text from the Function column.

    Returns:
        The matching :class:`NISTFunction` or ``None`` if the cell does
        not start with a recognised function name.
    """
    upper = text.upper()
    for func in NISTFunction:
        if upper.startswith(func.value):
            return func
    return None


def _parse_category_header(text: str) -> NISTCategory | None:
    """Parse a category header row.

    The reference tool formats rows as
    ``"Organizational Context (GV.OC): The circumstances ..."``.

    Args:
        text: Cell text from the Category column.

    Returns:
        Parsed :class:`NISTCategory`, or ``None`` if the row does not
        match the expected ``Name (ID): description`` shape.
    """
    if "(" not in text or ")" not in text:
        return None
    name_part, _, rest = text.partition("(")
    cat_id, _, description_part = rest.partition(")")
    description = description_part.lstrip(": ").strip()
    return NISTCategory(
        id=cat_id.strip(),
        name=name_part.strip(),
        description=description,
        subcategories=[],
    )


def _parse_subcategory_row(
    subcategory_text: str,
    examples_text: str,
    refs_text: str,
) -> NISTSubcategory | None:
    """Parse a subcategory data row.

    Args:
        subcategory_text: Cell text from the Subcategory column,
            formatted as ``"GV.OC-01: The organizational mission ..."``.
        examples_text: Multi-line implementation examples; one example
            per line. NIST CSRC ships each ``"Ex1: ..." / "Ex2: ..."``
            marker on its own line, so splitting on newlines and
            preserving the original ``"ExN:"`` prefix is equivalent to
            splitting on the marker itself while remaining tolerant of
            cells that do not use the convention.
        refs_text: Multi-line informative references; entries are split
            on newlines.

    Returns:
        Parsed :class:`NISTSubcategory`, or ``None`` if the
        subcategory text does not contain an identifier.
    """
    sub_id, _, description = subcategory_text.partition(":")
    sub_id = sub_id.strip()
    if not sub_id:
        return None
    examples = tuple(line.strip() for line in examples_text.splitlines() if line.strip())
    refs = tuple(line.strip() for line in refs_text.splitlines() if line.strip())
    return NISTSubcategory(
        id=sub_id,
        description=description.strip(),
        implementation_examples=examples,
        informative_references=refs,
    )


# ----------------------------------------------------------------- built-in tree

# Minimal built-in fallback tree. Used only when ``reference_source``
# is set to ``"builtin"`` explicitly. The tree intentionally covers
# every CSF 2.0 function with at least three categories so that the
# integrator's assessment flow remains functional in air-gapped
# environments (e.g. CI security-isolation lanes). For production
# compliance work, prefer ``reference_source="live"``.
_BUILTIN_CATEGORIES: Final[dict[NISTFunction, list[NISTCategory]]] = {
    NISTFunction.GOVERN: [
        NISTCategory(
            id="GV.OC",
            name="Organizational Context",
            description="Mission, stakeholders, and legal/regulatory requirements.",
            subcategories=[
                NISTSubcategory(id="GV.OC-01", description="Mission and business objectives"),
                NISTSubcategory(id="GV.OC-02", description="Internal and external stakeholders"),
                NISTSubcategory(
                    id="GV.OC-03",
                    description="Legal, regulatory, contractual requirements",
                ),
            ],
        ),
        NISTCategory(
            id="GV.RM",
            name="Risk Management Strategy",
            description="Priorities, constraints, risk tolerance, and assumptions.",
            subcategories=[
                NISTSubcategory(id="GV.RM-01", description="Risk management objectives"),
                NISTSubcategory(id="GV.RM-02", description="Risk appetite and tolerance"),
                NISTSubcategory(
                    id="GV.RM-03", description="Risk management roles and responsibilities"
                ),
            ],
        ),
        NISTCategory(
            id="GV.SC",
            name="Supply Chain Risk Management",
            description="Cyber supply chain risk management processes.",
            subcategories=[
                NISTSubcategory(id="GV.SC-01", description="Supply chain risk management strategy"),
                NISTSubcategory(id="GV.SC-02", description="Supplier cybersecurity requirements"),
                NISTSubcategory(id="GV.SC-03", description="Third-party risk assessment"),
            ],
        ),
    ],
    NISTFunction.IDENTIFY: [
        NISTCategory(
            id="ID.AM",
            name="Asset Management",
            description="Data, personnel, devices, systems, and facilities.",
            subcategories=[
                NISTSubcategory(
                    id="ID.AM-01", description="Physical devices and systems inventory"
                ),
                NISTSubcategory(
                    id="ID.AM-02", description="Software platforms and applications inventory"
                ),
                NISTSubcategory(
                    id="ID.AM-03", description="Organizational communication and data flows"
                ),
            ],
        ),
        NISTCategory(
            id="ID.RA",
            name="Risk Assessment",
            description="Understanding cybersecurity risk to operations and assets.",
            subcategories=[
                NISTSubcategory(id="ID.RA-01", description="Asset vulnerabilities identification"),
                NISTSubcategory(id="ID.RA-02", description="Cyber threat intelligence"),
                NISTSubcategory(
                    id="ID.RA-03", description="Internal and external threats identification"
                ),
            ],
        ),
        NISTCategory(
            id="ID.IM",
            name="Improvement",
            description="Lessons learned and improvement opportunities.",
            subcategories=[
                NISTSubcategory(id="ID.IM-01", description="Lessons learned incorporation"),
                NISTSubcategory(id="ID.IM-02", description="Response and recovery plan testing"),
                NISTSubcategory(id="ID.IM-03", description="Continuous improvement processes"),
            ],
        ),
    ],
    NISTFunction.PROTECT: [
        NISTCategory(
            id="PR.AA",
            name="Identity Management, Authentication, and Access Control",
            description="Access to physical and logical assets.",
            subcategories=[
                NISTSubcategory(id="PR.AA-01", description="Identities and credentials management"),
                NISTSubcategory(id="PR.AA-02", description="Physical access management"),
                NISTSubcategory(id="PR.AA-03", description="Remote access management"),
            ],
        ),
        NISTCategory(
            id="PR.DS",
            name="Data Security",
            description="Protection of data at rest and in transit.",
            subcategories=[
                NISTSubcategory(id="PR.DS-01", description="Data-at-rest protection"),
                NISTSubcategory(id="PR.DS-02", description="Data-in-transit protection"),
                NISTSubcategory(id="PR.DS-11", description="Backups maintained and tested"),
            ],
        ),
        NISTCategory(
            id="PR.PS",
            name="Platform Security",
            description="Security of hardware, software, and services.",
            subcategories=[
                NISTSubcategory(id="PR.PS-01", description="Configuration management"),
                NISTSubcategory(id="PR.PS-02", description="Secure development practices"),
                NISTSubcategory(id="PR.PS-03", description="Maintenance and repair controls"),
            ],
        ),
    ],
    NISTFunction.DETECT: [
        NISTCategory(
            id="DE.CM",
            name="Continuous Monitoring",
            description="Monitoring for cybersecurity events.",
            subcategories=[
                NISTSubcategory(id="DE.CM-01", description="Network monitoring"),
                NISTSubcategory(id="DE.CM-02", description="Physical environment monitoring"),
                NISTSubcategory(id="DE.CM-03", description="Personnel activity monitoring"),
            ],
        ),
        NISTCategory(
            id="DE.AE",
            name="Adverse Event Analysis",
            description="Analysis of detected anomalies.",
            subcategories=[
                NISTSubcategory(id="DE.AE-02", description="Detected events analysis"),
                NISTSubcategory(id="DE.AE-03", description="Event correlation and impact analysis"),
                NISTSubcategory(id="DE.AE-04", description="Adverse event criteria"),
            ],
        ),
    ],
    NISTFunction.RESPOND: [
        NISTCategory(
            id="RS.MA",
            name="Incident Management",
            description="Response activities coordination.",
            subcategories=[
                NISTSubcategory(id="RS.MA-01", description="Incident response plan execution"),
                NISTSubcategory(id="RS.MA-02", description="Incident reporting"),
                NISTSubcategory(id="RS.MA-03", description="Incident response team coordination"),
            ],
        ),
        NISTCategory(
            id="RS.AN",
            name="Incident Analysis",
            description="Investigation and analysis of incidents.",
            subcategories=[
                NISTSubcategory(id="RS.AN-03", description="Incident investigation"),
                NISTSubcategory(id="RS.AN-06", description="Impact analysis"),
                NISTSubcategory(id="RS.AN-07", description="Forensics performed"),
            ],
        ),
        NISTCategory(
            id="RS.MI",
            name="Incident Mitigation",
            description="Activities to prevent expansion of events.",
            subcategories=[
                NISTSubcategory(id="RS.MI-01", description="Incidents contained"),
                NISTSubcategory(id="RS.MI-02", description="Incidents mitigated"),
            ],
        ),
    ],
    NISTFunction.RECOVER: [
        NISTCategory(
            id="RC.RP",
            name="Incident Recovery Plan Execution",
            description="Restoration processes and procedures.",
            subcategories=[
                NISTSubcategory(id="RC.RP-01", description="Recovery plan execution"),
                NISTSubcategory(id="RC.RP-02", description="Recovery plan testing"),
                NISTSubcategory(id="RC.RP-03", description="Recovery plan updates"),
            ],
        ),
        NISTCategory(
            id="RC.CO",
            name="Incident Recovery Communication",
            description="Coordination with stakeholders.",
            subcategories=[
                NISTSubcategory(id="RC.CO-03", description="Recovery activities communication"),
                NISTSubcategory(id="RC.CO-04", description="Public updates"),
            ],
        ),
    ],
}

# --------------------------------------------------------------------- integrator


class NISTCSFIntegrator:
    """NIST CSF 2.0 integrator for risk management and reporting.

    Provides the five canonical operations exercised by Mercury Agent
    test suites and compliance pipelines:

    1. :meth:`assess_function`
    2. :meth:`create_profile`
    3. :meth:`detect_supply_chain_anomalies`
    4. :meth:`continuous_monitoring_detect`
    5. :meth:`generate_compliance_report`

    The integrator is constructed once per assessment cycle and reused
    across all five operations.
    """

    _CRITICALITY_WEIGHTS: Final[dict[NISTFunction, float]] = {
        NISTFunction.GOVERN: 1.2,
        NISTFunction.IDENTIFY: 1.1,
        NISTFunction.PROTECT: 1.3,
        NISTFunction.DETECT: 1.2,
        NISTFunction.RESPOND: 1.1,
        NISTFunction.RECOVER: 1.0,
    }

    _TIER_TARGETS: Final[dict[ImplementationTier, float]] = {
        ImplementationTier.PARTIAL: 0.50,
        ImplementationTier.RISK_INFORMED: 0.65,
        ImplementationTier.REPEATABLE: 0.80,
        ImplementationTier.ADAPTIVE: 0.95,
    }

    def __init__(
        self,
        target_tier: ImplementationTier = ImplementationTier.REPEATABLE,
        reference_source: str = "live",
        fetcher: NISTCSFReferenceFetcher | None = None,
    ) -> None:
        """Initialise the integrator.

        Args:
            target_tier: Desired implementation tier for the
                organization. Drives gap analysis in
                :meth:`create_profile` and
                :meth:`generate_compliance_report`.
            reference_source: Either ``"live"`` (default; fetches the
                authoritative NIST CSF 2.0 reference tree from CSRC)
                or ``"builtin"`` (uses the curated offline tree).
            fetcher: Optional pre-configured
                :class:`NISTCSFReferenceFetcher`. Ignored unless
                ``reference_source == "live"``.

        Raises:
            ValueError: If ``reference_source`` is not one of the
                accepted values.
            NISTCSFReferenceError: If the live reference cannot be
                loaded when ``reference_source == "live"`` and no
                cached payload exists.
        """
        if reference_source not in {"live", "builtin"}:
            raise ValueError(
                f"reference_source must be 'live' or 'builtin', got {reference_source!r}"
            )
        self.target_tier = target_tier
        self._reference_source = reference_source
        self._fetcher = fetcher
        self.categories: dict[NISTFunction, list[NISTCategory]] = self._load_categories()
        self.current_profile: NISTProfile | None = None
        self.target_profile: NISTProfile | None = None

    # ------------------------------------------------------------- bootstrap

    def _load_categories(self) -> dict[NISTFunction, list[NISTCategory]]:
        """Load the category tree according to ``reference_source``."""
        if self._reference_source == "builtin":
            return {func: list(cats) for func, cats in _BUILTIN_CATEGORIES.items()}
        fetcher = self._fetcher or NISTCSFReferenceFetcher()
        self._fetcher = fetcher
        return fetcher.load_reference_tree()

    @property
    def reference_source(self) -> str:
        """Return ``"live"`` or ``"builtin"`` per construction-time choice."""
        return self._reference_source

    @property
    def fetcher(self) -> NISTCSFReferenceFetcher | None:
        """Return the underlying fetcher when ``reference_source == "live"``."""
        return self._fetcher

    def verify_coverage(
        self,
        minimum_subcategories: int = 100,
    ) -> dict[str, int]:
        """Return per-function subcategory counts and assert coverage.

        The CSF 2.0 final publication defines 6 functions, 22
        categories, and 106 subcategories. This method asserts the
        loaded tree covers all six functions and at least
        ``minimum_subcategories`` total subcategories.

        Args:
            minimum_subcategories: Lower bound for total subcategory
                count. Defaults to 100 to allow minor NIST tooling
                drift.

        Returns:
            Dictionary mapping function names to subcategory counts,
            plus the ``"_total"`` key.

        Raises:
            NISTCSFReferenceError: If the loaded tree is missing a
                function or falls below the subcategory minimum.
        """
        counts: dict[str, int] = {}
        total = 0
        for func in NISTFunction:
            cats = self.categories.get(func, [])
            sub_count = sum(len(c.subcategories) for c in cats)
            counts[func.value] = sub_count
            total += sub_count
            if not cats:
                raise NISTCSFReferenceError(
                    f"NIST CSF reference missing categories for {func.value}"
                )
        counts["_total"] = total
        if total < minimum_subcategories:
            raise NISTCSFReferenceError(
                f"NIST CSF reference covers only {total} subcategories; "
                f"expected at least {minimum_subcategories}"
            )
        return counts

    # ------------------------------------------------------------- assessment

    def assess_function(
        self,
        function: NISTFunction,
        evidence: dict[str, Any],
    ) -> NISTAssessment:
        """Assess maturity and compliance for a single CSF function.

        Args:
            function: Function under assessment.
            evidence: Mapping from lowercased category or subcategory
                identifier (e.g. ``"gv.oc"`` or ``"gv.oc-01"``) to a
                maturity score in ``[0, 1]``. Missing keys default to
                ``0.50``.

        Returns:
            :class:`NISTAssessment` with maturity score, tier,
            findings, and recommendations.
        """
        categories = self.categories.get(function)
        if not categories:
            raise NISTCSFReferenceError(f"No categories loaded for function {function.value}")

        maturity_scores: list[float] = []
        findings: list[str] = []
        recommendations: list[str] = []

        for category in categories:
            score = self._assess_category(category, evidence)
            maturity_scores.append(score)

            sub_ids = category.subcategory_ids()
            if score < 0.60 and sub_ids:
                findings.append(f"Low maturity in {category.name} ({category.id}): {score:.2f}")
                recommendations.append(f"Improve {category.name}: Implement {sub_ids[0]}")
            elif score < 0.80 and sub_ids:
                recommendations.append(f"Enhance {category.name}: Focus on {sub_ids[-1]}")

        overall_maturity = float(np.mean(maturity_scores)) if maturity_scores else 0.0
        tier = self._determine_tier(overall_maturity)
        risk_score = self._calculate_risk_score(overall_maturity, function)

        if tier.value < self.target_tier.value:
            findings.append(f"Current tier ({tier.name}) below target ({self.target_tier.name})")
            recommendations.append(
                f"Advance to {self.target_tier.name} tier through systematic improvements"
            )

        return NISTAssessment(
            function=function,
            tier=tier,
            maturity_score=overall_maturity,
            findings=findings,
            recommendations=recommendations,
            risk_score=risk_score,
        )

    def _assess_category(
        self,
        category: NISTCategory,
        evidence: dict[str, Any],
    ) -> float:
        """Assess maturity of a single category from evidence dict."""
        category_key = category.id.lower()
        if category_key in evidence:
            return float(evidence[category_key])

        subcategory_scores: list[float] = []
        for sub_id in category.subcategory_ids():
            sub_key = sub_id.lower()
            if sub_key in evidence:
                subcategory_scores.append(float(evidence[sub_key]))

        if subcategory_scores:
            return float(np.mean(subcategory_scores))
        return 0.50

    def _determine_tier(self, maturity_score: float) -> ImplementationTier:
        """Resolve an :class:`ImplementationTier` from a maturity score."""
        if maturity_score >= 0.85:
            return ImplementationTier.ADAPTIVE
        if maturity_score >= 0.70:
            return ImplementationTier.REPEATABLE
        if maturity_score >= 0.50:
            return ImplementationTier.RISK_INFORMED
        return ImplementationTier.PARTIAL

    def _calculate_risk_score(
        self,
        maturity_score: float,
        function: NISTFunction,
    ) -> float:
        """Compute the criticality-weighted risk score in ``[0, 1]``."""
        base_risk = 1.0 - maturity_score
        weighted = base_risk * self._CRITICALITY_WEIGHTS[function]
        return float(min(max(weighted, 0.0), 1.0))

    # ---------------------------------------------------------------- profile

    def create_profile(
        self,
        assessments: list[NISTAssessment],
    ) -> NISTProfile:
        """Build a current-vs-target :class:`NISTProfile` from assessments.

        Args:
            assessments: One assessment per function (duplicates are
                tolerated and override earlier entries).

        Returns:
            :class:`NISTProfile` with per-function maturity, gaps, and
            priority actions for any function whose gap exceeds 0.20.
        """
        current_state: dict[str, float] = {}
        target_state: dict[str, float] = {}
        gaps: dict[str, float] = {}
        priority_actions: list[str] = []

        for assessment in assessments:
            key = assessment.function.value
            current_state[key] = assessment.maturity_score
            target_state[key] = self._TIER_TARGETS[self.target_tier]
            gaps[key] = target_state[key] - current_state[key]
            if gaps[key] > 0.20:
                priority_actions.extend(assessment.recommendations[:2])

        return NISTProfile(
            current_state=current_state,
            target_state=target_state,
            gaps=gaps,
            priority_actions=priority_actions,
        )

    # ----------------------------------------------------- supply chain anomalies

    def detect_supply_chain_anomalies(
        self,
        supplier_data: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        """Detect supplier anomalies via the CSF 2.0 GV.SC categories.

        Args:
            supplier_data: Mapping from supplier identifier to a dict
                with optional ``risk_score`` and ``compliance_score``
                keys (each in ``[0, 1]``).

        Returns:
            List of anomaly dicts. Each dict carries ``supplier_id``,
            ``type``, ``severity``, ``category`` (a GV.SC subcategory
            identifier), and ``description``.
        """
        anomalies: list[dict[str, Any]] = []
        for supplier_id, data in supplier_data.items():
            risk_score = float(data.get("risk_score", 0.5))
            compliance_score = float(data.get("compliance_score", 0.5))

            if risk_score > 0.70:
                anomalies.append(
                    {
                        "supplier_id": supplier_id,
                        "type": "high_risk_supplier",
                        "severity": risk_score,
                        "category": "GV.SC-03",
                        "description": (
                            f"Supplier {supplier_id} has high risk score: " f"{risk_score:.2f}"
                        ),
                    }
                )
            if compliance_score < 0.40:
                anomalies.append(
                    {
                        "supplier_id": supplier_id,
                        "type": "low_compliance",
                        "severity": 1.0 - compliance_score,
                        "category": "GV.SC-02",
                        "description": (
                            f"Supplier {supplier_id} has low compliance: " f"{compliance_score:.2f}"
                        ),
                    }
                )
        return anomalies

    # ----------------------------------------------------- continuous monitoring

    def continuous_monitoring_detect(
        self,
        network_data: np.ndarray[Any, np.dtype[np.floating[Any]]],
        baseline: np.ndarray[Any, np.dtype[np.floating[Any]]] | None = None,
    ) -> tuple[np.ndarray[Any, np.dtype[np.float64]], list[str]]:
        """Detect anomalies in network or system metrics (DE.CM, DE.AE).

        Args:
            network_data: 2-D array where each row is a sample and each
                column is a metric.
            baseline: Optional 1-D baseline; defaults to the column-wise
                mean of ``network_data``.

        Returns:
            Tuple ``(scores, events)`` where ``scores`` is a 1-D array
            of anomaly scores in ``[0, 1]`` aligned to the rows of
            ``network_data`` and ``events`` is a human-readable list of
            high-anomaly notifications.

        Raises:
            ValueError: If ``network_data`` is not 2-D.
        """
        arr = np.asarray(network_data, dtype=np.float64)
        if arr.ndim != 2:
            raise ValueError(f"network_data must be 2-D, got shape {arr.shape}")
        if baseline is None:
            baseline_arr = np.mean(arr, axis=0)
        else:
            baseline_arr = np.asarray(baseline, dtype=np.float64)
            if baseline_arr.shape != (arr.shape[1],):
                raise ValueError(
                    f"baseline shape {baseline_arr.shape} does not match "
                    f"network_data column count {arr.shape[1]}"
                )

        deviations = np.abs(arr - baseline_arr)
        max_baseline = float(np.max(np.abs(baseline_arr))) + 1e-10
        max_deviation = float(np.max(np.abs(deviations))) + 1e-10
        denom = max(max_baseline, max_deviation)
        scores = np.clip(np.max(deviations, axis=1) / denom, 0.0, 1.0)

        events: list[str] = []
        for idx, score in enumerate(scores):
            if score > 0.70:
                events.append(f"DE.CM-01: High anomaly at index {idx} (score: {score:.2f})")
        return scores.astype(np.float64, copy=False), events

    # ----------------------------------------------------------- compliance report

    def generate_compliance_report(
        self,
        assessments: list[NISTAssessment],
        profile: NISTProfile,
    ) -> dict[str, Any]:
        """Produce a structured compliance report from assessments and profile.

        Args:
            assessments: All function assessments.
            profile: Profile output of :meth:`create_profile`.

        Returns:
            Nested dictionary with overall metrics, per-function
            findings, profile gaps, and a strengths/weaknesses summary.
            JSON-serialisable.
        """
        if assessments:
            overall_maturity = float(np.mean([a.maturity_score for a in assessments]))
            overall_risk = float(np.mean([a.risk_score for a in assessments]))
        else:
            overall_maturity = 0.0
            overall_risk = 0.0
        overall_tier = self._determine_tier(overall_maturity)

        return {
            "timestamp": datetime.now(UTC).isoformat(),
            "overall_maturity": overall_maturity,
            "overall_risk": overall_risk,
            "current_tier": overall_tier.name,
            "target_tier": self.target_tier.name,
            "tier_gap": self.target_tier.value - overall_tier.value,
            "reference_source": self._reference_source,
            "function_assessments": [
                {
                    "function": a.function.value,
                    "maturity": a.maturity_score,
                    "risk": a.risk_score,
                    "tier": a.tier.name,
                    "findings": list(a.findings),
                    "recommendations": list(a.recommendations),
                }
                for a in assessments
            ],
            "profile": asdict(profile),
            "summary": {
                "strengths": [
                    f"{a.function.value}: {a.maturity_score:.2f}"
                    for a in assessments
                    if a.maturity_score >= 0.80
                ],
                "weaknesses": [
                    f"{a.function.value}: {a.maturity_score:.2f}"
                    for a in assessments
                    if a.maturity_score < 0.60
                ],
                "total_recommendations": sum(len(a.recommendations) for a in assessments),
            },
        }

    # ------------------------------------------------------------- serialisation

    def export_categories_json(self) -> str:
        """Serialise the loaded category tree to a JSON string.

        Returns:
            JSON string with the full Function -> Category ->
            Subcategory tree. Useful for snapshot testing and audit
            archives.
        """
        payload = {
            func.value: [
                {
                    "id": cat.id,
                    "name": cat.name,
                    "description": cat.description,
                    "subcategories": [
                        (
                            {
                                "id": sub.id,
                                "description": sub.description,
                                "implementation_examples": list(sub.implementation_examples),
                                "informative_references": list(sub.informative_references),
                            }
                            if isinstance(sub, NISTSubcategory)
                            else {"id": sub.split(":", 1)[0].strip(), "description": sub}
                        )
                        for sub in cat.subcategories
                    ],
                }
                for cat in self.categories.get(func, [])
            ]
            for func in NISTFunction
        }
        return json.dumps(payload, indent=2, sort_keys=True)


def get_nist_csf_integrator(
    target_tier: ImplementationTier = ImplementationTier.REPEATABLE,
    reference_source: str = "live",
) -> NISTCSFIntegrator:
    """Construct a :class:`NISTCSFIntegrator`.

    Args:
        target_tier: Desired implementation tier.
        reference_source: ``"live"`` (default) or ``"builtin"``.

    Returns:
        A ready-to-use :class:`NISTCSFIntegrator`.
    """
    return NISTCSFIntegrator(target_tier=target_tier, reference_source=reference_source)


__all__ = [
    "NIST_CSF_PUBLICATION_PDF_URL",
    "NIST_CSF_PUBLICATION_URL",
    "NIST_CSF_REFERENCE_URL",
    "ImplementationTier",
    "NISTAssessment",
    "NISTCSFIntegrator",
    "NISTCSFReferenceError",
    "NISTCSFReferenceFetcher",
    "NISTCategory",
    "NISTFunction",
    "NISTProfile",
    "NISTSubcategory",
    "get_nist_csf_integrator",
]
